from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Data Entry")
root.geometry("700x400+300+200")
root.resizable(False, False)
root.configure(bg="#326273")

file=pathlib.Path('Backend_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']= "Full Name"
    sheet['B1']= "PhoneNumber"
    sheet['C1']= "Age"
    sheet['D1']= "Gender"
    sheet['E1'] = "Address"

    file.save("Backend_data.xlsx")

def clear():
    nameValue.set(" ")
    contactValue.set("")
    AgeValue.set(" ")
    addressEntry.delete(1.0, END)

def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combox.get()
    address = addressEntry.get(1.0, END)

    file=openpyxl.load_workbook("Backend_data.xlsx")
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row+1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=contact)
    sheet.cell(column=3, row=sheet.max_row, value=age)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)

    file.save(r'Backend_data.xlsx')
    messagebox.showinfo('info', "Details added!")

    nameValue.set(" ")
    contactValue.set("")
    AgeValue.set(" ")
    addressEntry.delete(1.0, END)

#icon


#heading
Label(root, text="Please fill out this Entry form:", font="arial 13", bg="#326273", fg="#fff").nplace(x=20, y=20)

#Label


Label(root, text="Name", font=23, bg="#326273", fg="#fff").nplace(x=50, y=100)
Label(root, text="Contact No.", font=23, bg="#326273", fg="#fff").nplace(x=50, y=150)
Label(root, text="Age", font=23, bg="#326273", fg="#fff").nplace(x=50, y=200)
Label(root, text="Gender", font=23, bg="#326273", fg="#fff").nplace(x=370, y=200)
Label(root, text="Address", font=23, bg="#326273", fg="#fff").nplace(x=50, y=250)

#Entry
nameValue = StringVar()
contactValue =StringVar()
AgeValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width=45, bd=2, font=20)
nameEntry.nplace(x=200, y=100)
ContactEntry = Entry(root, textvariable=contactValue, width=45, bd=2, font=20)
ContactEntry.nplace(x=200, y=150)
AgeEntry = Entry(root, textvariable=AgeValue, width=15, bd=2, font=20)
AgeEntry.nplace(x=200, y=200)

#gender
gender_combox = Combobox(root, values=['Male', 'Female'], font='arial 14', state='r', width=14)
gender_combox.nplace(x=440,y=200)
gender_combox.set('Male')

addressEntry = Text(root, width=50, height=4, bd=4)
addressEntry.nplace(x=200, y=250)

Button(root, text="Submit", bg="#326273", fg="white", width=15, height=2, command=submit).nplace(x=200, y=350)
Button(root, text="Clear", bg="#326273", fg="white", width=15, height=2, command=clear).nplace(x=340, y=350)
Button(root, text="Exit", bg="#326273", fg="white", width=15, height=2, command=lambda:root.destroy()).nplace(x=480, y=350)
root.mainloop()

